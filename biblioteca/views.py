from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.views import LoginView as AuthLoginView, LogoutView as AuthLogoutView
from django.contrib.auth import login
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.urls import reverse_lazy
from django.core.exceptions import ValidationError
from django.utils import timezone
from .forms import LoginForm, RegisterForm, LivroForm, AutorForm, CategoriaForm, EmprestimoForm, ReservaForm, ProfileForm
from .models import Livro, Autor, Categoria, Emprestimo, Reserva, Usuario
import datetime
from django.db import models

# Mixin for admin-only access
class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_admin()

# Basic views
class HomeView(AdminRequiredMixin, TemplateView):
    template_name = 'biblioteca/home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_livros'] = Livro.objects.count()
        context['total_usuarios'] = Usuario.objects.count()
        context['emprestimos_ativos'] = Emprestimo.objects.filter(status='ativa').count()
        context['reservas_ativas'] = Reserva.objects.filter(status='ativa').count()
        return context

class LoginView(AuthLoginView):
    template_name = 'biblioteca/login.html'
    form_class = LoginForm
    
    def get_success_url(self):
        if self.request.user.is_admin():
            return reverse_lazy('biblioteca:admin_dashboard')
        return reverse_lazy('biblioteca:home')

class LogoutView(AuthLogoutView):
    next_page = 'biblioteca:home'

class RegisterView(TemplateView):
    template_name = 'biblioteca/register.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RegisterForm()
        return context
    
    def post(self, request, *args, **kwargs):
        form = RegisterForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                # Definir tipo de usuário padrão como aluno
                user.tipo_usuario = 'aluno'
                user.save()
                
                # Fazer login automático após registro
                login(request, user)
                messages.success(request, 'Conta criada com sucesso! Bem-vindo à Biblioteca SENAC!')
                return redirect('biblioteca:dashboard')
            except Exception as e:
                messages.error(request, f'Erro ao criar conta: {str(e)}')
        else:
            # Exibir erros de validação
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
        
        context = self.get_context_data()
        context['form'] = form
        return render(request, self.template_name, context)

class ProfileView(LoginRequiredMixin, TemplateView):
    template_name = 'biblioteca/profile.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ProfileForm(instance=self.request.user)
        context['user'] = self.request.user
        
        # Estatísticas do usuário
        context['total_reservas'] = self.request.user.get_total_reservations()
        context['reservas_ativas'] = self.request.user.get_active_reservations()
        context['total_emprestimos'] = self.request.user.get_total_loans()
        context['emprestimos_ativos'] = self.request.user.get_active_loans()
        
        return context
    
    def post(self, request, *args, **kwargs):
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('biblioteca:profile')
        else:
            messages.error(request, 'Erro ao atualizar perfil. Verifique os dados.')
        
        context = self.get_context_data()
        context['form'] = form
        return render(request, self.template_name, context)

# Livro views
class LivroListView(ListView):
    model = Livro
    template_name = 'biblioteca/livro_list.html'
    context_object_name = 'livros'
    paginate_by = 12
    
    def get_queryset(self):
        queryset = Livro.objects.select_related('autor').all()
        
        # Search functionality
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(titulo__icontains=query) |
                Q(autor__nome__icontains=query)
            )
        
        # Filter by genre
        genero = self.request.GET.get('genero')
        if genero:
            queryset = queryset.filter(genero=genero)
            
        return queryset.order_by('titulo')

class LivroDetailView(DetailView):
    model = Livro
    template_name = 'biblioteca/livro_detail.html'
    context_object_name = 'livro'

class LivroBuscarView(TemplateView):
    template_name = 'biblioteca/livro_buscar.html'

class LivroCreateView(AdminRequiredMixin, CreateView):
    model = Livro
    form_class = LivroForm
    template_name = 'biblioteca/livro_form.html'
    success_url = reverse_lazy('biblioteca:livro_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Livro criado com sucesso!')
        return super().form_valid(form)

class LivroUpdateView(AdminRequiredMixin, UpdateView):
    model = Livro
    form_class = LivroForm
    template_name = 'biblioteca/livro_form.html'
    success_url = reverse_lazy('biblioteca:livro_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Livro atualizado com sucesso!')
        return super().form_valid(form)

class LivroDeleteView(AdminRequiredMixin, DeleteView):
    model = Livro
    template_name = 'biblioteca/livro_confirm_delete.html'
    success_url = reverse_lazy('biblioteca:livro_list')

# Autor views
class AutorListView(ListView):
    model = Autor
    template_name = 'biblioteca/autor_list.html'
    context_object_name = 'autores'

class AutorDetailView(DetailView):
    model = Autor
    template_name = 'biblioteca/autor_detail.html'
    context_object_name = 'autor'

class AutorCreateView(AdminRequiredMixin, CreateView):
    model = Autor
    form_class = AutorForm
    template_name = 'biblioteca/autor_form.html'
    success_url = reverse_lazy('biblioteca:autor_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Autor criado com sucesso!')
        return super().form_valid(form)

class AutorUpdateView(AdminRequiredMixin, UpdateView):
    model = Autor
    form_class = AutorForm
    template_name = 'biblioteca/autor_form.html'
    success_url = reverse_lazy('biblioteca:autor_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Autor atualizado com sucesso!')
        return super().form_valid(form)

class AutorDeleteView(AdminRequiredMixin, DeleteView):
    model = Autor
    template_name = 'biblioteca/autor_confirm_delete.html'
    success_url = reverse_lazy('biblioteca:autor_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Autor excluído com sucesso!')
        return super().delete(request, *args, **kwargs)

# Reserva views
class ReservaListView(AdminRequiredMixin, ListView):
    model = Reserva
    template_name = 'biblioteca/reserva_list.html'
    context_object_name = 'reservas'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Reserva.objects.select_related('usuario', 'livro', 'livro__autor').all()
        
        # Search functionality
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(usuario__username__icontains=query) |
                Q(usuario__first_name__icontains=query) |
                Q(usuario__last_name__icontains=query) |
                Q(usuario__email__icontains=query) |
                Q(livro__titulo__icontains=query) |
                Q(livro__autor__nome__icontains=query) |
                Q(status__icontains=query)
            )
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        # Filter by user type
        tipo_usuario = self.request.GET.get('tipo_usuario')
        if tipo_usuario:
            queryset = queryset.filter(usuario__tipo_usuario=tipo_usuario)
            
        # Filter by date
        data_inicio = self.request.GET.get('data_inicio')
        if data_inicio:
            queryset = queryset.filter(data_reserva__date__gte=data_inicio)
            
        return queryset.order_by('-data_reserva')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add current time for expiration checking
        context['now'] = timezone.now()
        
        # Add statistics
        queryset = self.get_queryset()
        context['total_reservas'] = queryset.count()
        context['reservas_ativas'] = queryset.filter(status='ativa').count()
        context['reservas_expiradas'] = queryset.filter(status='expirada').count()
        context['reservas_canceladas'] = queryset.filter(status='cancelada').count()
        
        return context

class MinhasReservasView(LoginRequiredMixin, ListView):
    model = Reserva
    template_name = 'biblioteca/minhas_reservas.html'
    context_object_name = 'reservas'

    def get_queryset(self):
        # This is not used by the template, so we override get_context_data
        return Reserva.objects.filter(usuario=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_reservas = Reserva.objects.filter(usuario=self.request.user)
        context['reservas_ativas'] = user_reservas.filter(status='ativa')
        context['reservas_historico'] = user_reservas.exclude(status='ativa')
        context['reservas_expiradas'] = user_reservas.filter(status='expirada')
        context['reservas_canceladas'] = user_reservas.filter(status='cancelada')
        context['total_reservas'] = user_reservas.count()
        return context

class ReservarLivroView(LoginRequiredMixin, CreateView):
    model = Reserva
    template_name = 'biblioteca/reservar_livro.html'
    fields = []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['livro'] = get_object_or_404(Livro, pk=self.kwargs['livro_id'])
        return context
    
    def post(self, request, *args, **kwargs):
        livro = get_object_or_404(Livro, pk=self.kwargs['livro_id'])
        
        # Verificar se o livro está disponível
        if livro.quantidade_disponivel <= 0:
            messages.error(request, 'Este livro não está disponível para reserva.')
            return render(request, self.template_name, {
                'livro': livro,
                'form': self.get_form()
            })
        
        # Verificar se o usuário pode fazer mais reservas
        if not request.user.pode_reservar():
            messages.error(request, 'Você já possui o limite máximo de reservas ativas (3).')
            return render(request, self.template_name, {
                'livro': livro,
                'form': self.get_form()
            })
        
        # Create the reservation instance manually
        reserva = Reserva(
            usuario=request.user,
            livro=livro,
            status='ativa'
        )
        
        try:
            # Validate the instance
            reserva.clean()
            reserva.save()
            
            # Recalcular quantidade disponível do livro
            livro.recalcular_quantidade_disponivel()
            
            messages.success(request, 'Reserva realizada com sucesso!')
            return redirect('biblioteca:minhas_reservas')
        except ValidationError as e:
            messages.error(request, str(e))
            return render(request, self.template_name, {
                'livro': livro,
                'form': self.get_form()
            })
    
    def get_success_url(self):
        return reverse_lazy('biblioteca:minhas_reservas')

class CancelarReservaView(LoginRequiredMixin, UpdateView):
    model = Reserva
    fields = []
    http_method_names = ['post']  # Apenas aceita POST
    
    def post(self, request, *args, **kwargs):
        try:
            reserva = self.get_object()
            
            # Verificar permissões
            if reserva.usuario != request.user and not request.user.is_admin():
                messages.error(request, 'Você não tem permissão para cancelar esta reserva.')
                return redirect('biblioteca:minhas_reservas')
            
            # Verificar se a reserva pode ser cancelada
            if reserva.status != 'ativa':
                messages.warning(request, f'Esta reserva não pode ser cancelada (status: {reserva.get_status_display()}).')
                return redirect('biblioteca:minhas_reservas')
            
            # Cancelar a reserva
            reserva.status = 'cancelada'
            reserva.save()
            
            # Atualizar quantidade disponível do livro
            livro = reserva.livro
            livro.recalcular_quantidade_disponivel()
            
            # Log da ação
            messages.success(request, f'Reserva do livro "{reserva.livro.titulo}" cancelada com sucesso!')
            
            # Redirecionar baseado no usuário
            if request.user.is_admin():
                return redirect('biblioteca:admin_dashboard')
            else:
                return redirect('biblioteca:minhas_reservas')
                
        except Reserva.DoesNotExist:
            messages.error(request, 'Reserva não encontrada.')
            return redirect('biblioteca:minhas_reservas')
        except Exception as e:
            messages.error(request, f'Erro ao cancelar reserva: {str(e)}')
            return redirect('biblioteca:minhas_reservas')

class ReservaDetailView(DetailView):
    model = Reserva
    template_name = 'biblioteca/reserva_detail.html'
    context_object_name = 'reserva'

# Admin views
class AdminDashboardView(AdminRequiredMixin, TemplateView):
    template_name = 'biblioteca/admin_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_livros'] = Livro.objects.count()
        context['total_usuarios'] = Usuario.objects.count()
        context['emprestimos_ativos'] = Emprestimo.objects.filter(status='ativa').count()
        context['reservas_ativas'] = Reserva.objects.filter(status='ativa').count()
        # Optionally add more context for alerts and recent activities
        context['emprestimos_vencidos'] = Emprestimo.objects.filter(status='vencido')
        context['reservas_expirando'] = Reserva.objects.filter(status='expirada')
        context['livros_sem_estoque'] = Livro.objects.filter(quantidade_disponivel=0)
        context['atividades_recentes'] = []  # Add your logic for recent activities here
        return context

class UsuarioListView(AdminRequiredMixin, ListView):
    model = Usuario
    template_name = 'biblioteca/usuario_list.html'
    context_object_name = 'usuarios'
    paginate_by = 12

    def get_queryset(self):
        queryset = Usuario.objects.all()
        search = self.request.GET.get('search')
        tipo_usuario = self.request.GET.get('tipo_usuario')
        status = self.request.GET.get('status')
        ordenar = self.request.GET.get('ordenar')

        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        if tipo_usuario:
            queryset = queryset.filter(tipo_usuario=tipo_usuario)
        if status:
            if status == 'ativo':
                queryset = queryset.filter(is_active=True)
            elif status == 'inativo':
                queryset = queryset.filter(is_active=False)
        if ordenar:
            if ordenar == 'nome':
                queryset = queryset.order_by('first_name', 'last_name')
            elif ordenar == 'data_cadastro':
                queryset = queryset.order_by('-date_joined')
            elif ordenar == 'ultimo_acesso':
                queryset = queryset.order_by('-last_login')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuarios = context['usuarios']

        # Estatísticas
        now = datetime.datetime.now()
        context['total_usuarios'] = Usuario.objects.count()
        context['usuarios_ativos'] = Usuario.objects.filter(is_active=True).count()
        context['usuarios_admin'] = Usuario.objects.filter(tipo_usuario='admin').count()
        context['novos_usuarios_mes'] = Usuario.objects.filter(
            date_joined__month=now.month,
            date_joined__year=now.year
        ).count()

        # Estatísticas individuais
        for usuario in usuarios:
            usuario.emprestimos_count = Emprestimo.objects.filter(usuario=usuario).count()
            usuario.reservas_count = Reserva.objects.filter(usuario=usuario).count()
            usuario.pendencias_count = Emprestimo.objects.filter(usuario=usuario, status='vencido').count()

        return context

class UsuarioDetailView(AdminRequiredMixin, DetailView):
    model = Usuario
    template_name = 'biblioteca/usuario_detail.html'
    context_object_name = 'usuario'

class UsuarioUpdateView(AdminRequiredMixin, UpdateView):
    model = Usuario
    template_name = 'biblioteca/usuario_form.html'
    fields = ['first_name', 'last_name', 'email', 'tipo_usuario', 'is_active']
    success_url = reverse_lazy('biblioteca:usuario_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Usuário atualizado com sucesso!')
        return super().form_valid(form)

class UsuarioDeleteView(AdminRequiredMixin, DeleteView):
    model = Usuario
    template_name = 'biblioteca/usuario_confirm_delete.html'
    success_url = reverse_lazy('biblioteca:usuario_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Usuário excluído com sucesso!')
        return super().delete(request, *args, **kwargs)

# Emprestimo views
class EmprestimoListView(AdminRequiredMixin, ListView):
    model = Emprestimo
    template_name = 'biblioteca/emprestimo_list.html'
    context_object_name = 'emprestimos'

class EmprestimoDetailView(DetailView):
    model = Emprestimo
    template_name = 'biblioteca/emprestimo_detail.html'
    context_object_name = 'emprestimo'

class EmprestimoCreateView(AdminRequiredMixin, CreateView):
    model = Emprestimo
    form_class = EmprestimoForm
    template_name = 'biblioteca/emprestimo_form.html'
    success_url = reverse_lazy('biblioteca:emprestimo_list')
    
    def get_initial(self):
        initial = super().get_initial()
        # Se um livro foi especificado na URL, pré-seleciona ele
        livro_id = self.request.GET.get('livro')
        if livro_id:
            try:
                livro = Livro.objects.get(pk=livro_id)
                initial['livro'] = livro
            except Livro.DoesNotExist:
                pass
        return initial
    
    def form_valid(self, form):
        try:
            # Verificar se o livro está disponível
            livro = form.instance.livro
            if livro.quantidade_disponivel > 0:
                # Salvar o empréstimo primeiro
                emprestimo = form.save()
                
                # Recalcular quantidade disponível do livro
                livro.recalcular_quantidade_disponivel()
                
                messages.success(self.request, f'Empréstimo criado com sucesso! O livro "{livro.titulo}" foi emprestado para {form.instance.usuario.get_full_name() or form.instance.usuario.username}.')
                return redirect('biblioteca:emprestimo_list')
            else:
                form.add_error('livro', 'Este livro não está disponível para empréstimo.')
                messages.error(self.request, 'Livro não disponível para empréstimo.')
                return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'Erro ao criar empréstimo: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, corrija os erros abaixo.')
        return super().form_invalid(form)

# Categoria views
class CategoriaListView(ListView):
    model = Categoria
    template_name = 'biblioteca/categoria_list.html'
    context_object_name = 'categorias'

class CategoriaDetailView(DetailView):
    model = Categoria
    template_name = 'biblioteca/categoria_detail.html'
    context_object_name = 'categoria'

class CategoriaCreateView(AdminRequiredMixin, CreateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'biblioteca/categoria_form.html'
    success_url = reverse_lazy('biblioteca:categoria_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Categoria criada com sucesso!')
        return super().form_valid(form)

class CategoriaUpdateView(AdminRequiredMixin, UpdateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'biblioteca/categoria_form.html'
    success_url = reverse_lazy('biblioteca:categoria_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Categoria atualizada com sucesso!')
        return super().form_valid(form)

class CategoriaDeleteView(AdminRequiredMixin, DeleteView):
    model = Categoria
    template_name = 'biblioteca/categoria_confirm_delete.html'
    success_url = reverse_lazy('biblioteca:categoria_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Categoria excluída com sucesso!')
        return super().delete(request, *args, **kwargs)

# Dashboard views
class RelatoriosView(AdminRequiredMixin, TemplateView):
    template_name = 'biblioteca/relatorios.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data_inicio = self.request.GET.get('data_inicio')
        data_fim = self.request.GET.get('data_fim')
        tipo_relatorio = self.request.GET.get('tipo_relatorio')

        emprestimos = Emprestimo.objects.all()
        if data_inicio:
            emprestimos = emprestimos.filter(data_emprestimo__gte=data_inicio)
        if data_fim:
            emprestimos = emprestimos.filter(data_emprestimo__lte=data_fim)

        total_emprestimos = emprestimos.count()

        # Livros mais emprestados (correto)
        livros_populares = (
            Livro.objects
            .filter(id__in=emprestimos.values_list('livro', flat=True))
            .annotate(total_emprestimos=models.Count('emprestimos'))
            .order_by('-total_emprestimos')[:10]
        )

        usuarios_ativos = Usuario.objects.filter(
            id__in=emprestimos.filter(status='ativa').values_list('usuario', flat=True)
        ).count()

        devolvidos_no_prazo = emprestimos.filter(
            status='devolvido',
            data_devolucao__lte=models.F('data_devolucao_prevista')
        ).count()
        total_devolvidos = emprestimos.filter(status='devolvido').count()
        taxa_devolucao = (
            (devolvidos_no_prazo / total_devolvidos) * 100 if total_devolvidos > 0 else 0
        )

        context['stats'] = {
            'total_emprestimos': total_emprestimos,
            'livros_populares': livros_populares.count(),
            'usuarios_ativos': usuarios_ativos,
            'taxa_devolucao': round(taxa_devolucao, 2),
        }
        context['emprestimos'] = emprestimos
        context['livros_populares'] = livros_populares
        context['today'] = datetime.date.today()
        return context

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'biblioteca/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Estatísticas principais
        context['stats'] = {
            'total_livros': Livro.objects.count(),
            'total_usuarios': Usuario.objects.filter(is_active=True).count(),
            'emprestimos_ativos': Emprestimo.objects.filter(status='ativa').count(),
            'emprestimos_atrasados': Emprestimo.objects.filter(status='vencido').count(),
        }
        # Livros mais emprestados
        context['livros_populares'] = (
            Livro.objects
            .annotate(total_emprestimos=models.Count('emprestimos'))
            .order_by('-total_emprestimos')[:10]
        )
        return context

class VerificarDisponibilidadeView(TemplateView):
    template_name = 'biblioteca/verificar_disponibilidade.html'

class ExpirarReservasView(AdminRequiredMixin, TemplateView):
    template_name = 'biblioteca/expirar_reservas.html'

# Function-based views
def devolver_livro(request, pk):
    """
    View para devolver um livro emprestado.
    Apenas admins ou o próprio usuário podem devolver o livro.
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False, 
            'error': 'Método não permitido. Use POST.'
        }, status=405)
    
    try:
        emprestimo = get_object_or_404(Emprestimo, pk=pk)
        
        # Verificar se o usuário tem permissão (admin ou próprio usuário)
        if not (request.user.is_authenticated and 
                (request.user.is_admin() or request.user == emprestimo.usuario)):
            return JsonResponse({
                'success': False, 
                'error': 'Você não tem permissão para devolver este livro.'
            }, status=403)
        
        # Verificar se o empréstimo já foi devolvido
        if emprestimo.status == 'devolvido':
            return JsonResponse({
                'success': False, 
                'error': 'Este empréstimo já foi devolvido.'
            }, status=400)
        
        # Verificar se o empréstimo está ativo
        if emprestimo.status != 'ativo':
            return JsonResponse({
                'success': False, 
                'error': f'Não é possível devolver um empréstimo com status "{emprestimo.get_status_display()}".'
            }, status=400)
        
        # Tentar devolver
        if emprestimo.devolver():
            # Log da devolução
            print(f"Empréstimo {emprestimo.pk} devolvido por {request.user.username} em {timezone.now()}")
            
            messages.success(request, 'Livro devolvido com sucesso!')
            
            return JsonResponse({
                'success': True,
                'message': 'Livro devolvido com sucesso!',
                'data_devolucao': emprestimo.data_devolucao.strftime("%d/%m/%Y %H:%M") if emprestimo.data_devolucao else None
            })
        else:
            return JsonResponse({
                'success': False, 
                'error': 'Erro interno ao devolver o livro. Tente novamente.'
            }, status=500)
        
    except Emprestimo.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Empréstimo não encontrado.'
        }, status=404)
    except Exception as e:
        print(f"Erro ao devolver livro {pk}: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'Erro interno do servidor. Tente novamente.'
        }, status=500)

def renovar_emprestimo(request, pk):
    """
    View para renovar um empréstimo.
    Apenas empréstimos ativos, não atrasados e com menos de 2 renovações podem ser renovados.
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False, 
            'error': 'Método não permitido. Use POST.'
        }, status=405)
    
    try:
        emprestimo = get_object_or_404(Emprestimo, pk=pk)
        
        # Verificar se o usuário tem permissão (admin ou próprio usuário)
        if not (request.user.is_authenticated and 
                (request.user.is_admin() or request.user == emprestimo.usuario)):
            return JsonResponse({
                'success': False, 
                'error': 'Você não tem permissão para renovar este empréstimo.'
            }, status=403)
        
        # Validações antes de renovar
        if emprestimo.status != 'ativo':
            return JsonResponse({
                'success': False, 
                'error': f'Não é possível renovar um empréstimo com status "{emprestimo.get_status_display()}".'
            })
        
        if emprestimo.is_atrasado():
            return JsonResponse({
                'success': False, 
                'error': 'Não é possível renovar um empréstimo em atraso.'
            })
        
        if emprestimo.renovacoes >= 2:
            return JsonResponse({
                'success': False, 
                'error': 'Este empréstimo já atingiu o limite máximo de 2 renovações.'
            })
        
        # Tentar renovar
        if emprestimo.renovar():
            # Log da renovação
            print(f"Empréstimo {emprestimo.pk} renovado por {request.user.username} em {timezone.now()}")
            
            messages.success(
                request, 
                f'Empréstimo renovado com sucesso! Nova data de devolução: {emprestimo.data_devolucao_prevista.strftime("%d/%m/%Y")}'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Empréstimo renovado com sucesso!',
                'nova_data_devolucao': emprestimo.data_devolucao_prevista.strftime("%d/%m/%Y"),
                'renovacoes_restantes': 2 - emprestimo.renovacoes
            })
        else:
            return JsonResponse({
                'success': False, 
                'error': 'Erro interno ao renovar o empréstimo.'
            })
            
    except Emprestimo.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Empréstimo não encontrado.'
        }, status=404)
    except Exception as e:
        print(f"Erro ao renovar empréstimo {pk}: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'Erro interno do servidor. Tente novamente.'
        }, status=500)

def livros_disponiveis(request):
    livros = Livro.objects.filter(quantidade_disponivel__gt=0)
    data = [{'id': l.id, 'titulo': l.titulo, 'autor': l.autor.nome} for l in livros]
    return JsonResponse({'livros': data})

def verificar_disponibilidade(request, livro_id):
    livro = get_object_or_404(Livro, pk=livro_id)
    return JsonResponse({
        'disponivel': livro.is_available(),
        'quantidade_disponivel': livro.quantidade_disponivel,
        'quantidade_total': livro.quantidade
    })

# View para deletar reserva (admin only)
class DeletarReservaView(AdminRequiredMixin, DeleteView):
    model = Reserva
    template_name = 'biblioteca/reserva_confirm_delete.html'
    success_url = reverse_lazy('biblioteca:reserva_list')
    
    def delete(self, request, *args, **kwargs):
        reserva = self.get_object()
        
        # Se a reserva estiver ativa, liberar o livro
        if reserva.status == 'ativa':
            livro = reserva.livro
            if livro.quantidade_disponivel < livro.quantidade:
                livro.quantidade_disponivel += 1
                livro.save()
        
        messages.success(request, f'Reserva deletada com sucesso!')
        return super().delete(request, *args, **kwargs)

# View para exportar reservas
class ExportarReservasView(AdminRequiredMixin, TemplateView):
    template_name = 'biblioteca/exportar_reservas.html'
    
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'csv')
        
        # Aplicar os mesmos filtros da listagem
        queryset = Reserva.objects.select_related('usuario', 'livro', 'livro__autor').all()
        
        query = request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(usuario__username__icontains=query) |
                Q(usuario__first_name__icontains=query) |
                Q(usuario__last_name__icontains=query) |
                Q(usuario__email__icontains=query) |
                Q(livro__titulo__icontains=query) |
                Q(livro__autor__nome__icontains=query) |
                Q(status__icontains=query)
            )
        
        status = request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        tipo_usuario = request.GET.get('tipo_usuario')
        if tipo_usuario:
            queryset = queryset.filter(usuario__tipo_usuario=tipo_usuario)
            
        data_inicio = request.GET.get('data_inicio')
        if data_inicio:
            queryset = queryset.filter(data_reserva__date__gte=data_inicio)
        
        if format_type == 'csv':
            return self.export_csv(queryset)
        elif format_type == 'excel':
            return self.export_excel(queryset)
        elif format_type == 'pdf':
            return self.export_pdf(queryset)
        else:
            return self.export_csv(queryset)
    
    def export_csv(self, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reservas.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Usuário', 'Email', 'Tipo Usuário', 'Livro', 'Autor', 
            'Status', 'Data Reserva', 'Data Expiração'
        ])
        
        for reserva in queryset:
            writer.writerow([
                reserva.id,
                reserva.usuario.get_full_name() or reserva.usuario.username,
                reserva.usuario.email,
                reserva.usuario.get_tipo_usuario_display(),
                reserva.livro.titulo,
                reserva.livro.autor.nome,
                reserva.get_status_display(),
                reserva.data_reserva.strftime('%d/%m/%Y %H:%M'),
                reserva.data_expiracao.strftime('%d/%m/%Y %H:%M') if reserva.data_expiracao else 'N/A'
            ])
        
        return response
    
    def export_excel(self, queryset):
        try:
            import openpyxl
            from openpyxl import Workbook
            from django.http import HttpResponse
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reservas"
            
            # Headers
            headers = [
                'ID', 'Usuário', 'Email', 'Tipo Usuário', 'Livro', 'Autor', 
                'Status', 'Data Reserva', 'Data Expiração'
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row, reserva in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=reserva.id)
                ws.cell(row=row, column=2, value=reserva.usuario.get_full_name() or reserva.usuario.username)
                ws.cell(row=row, column=3, value=reserva.usuario.email)
                ws.cell(row=row, column=4, value=reserva.usuario.get_tipo_usuario_display())
                ws.cell(row=row, column=5, value=reserva.livro.titulo)
                ws.cell(row=row, column=6, value=reserva.livro.autor.nome)
                ws.cell(row=row, column=7, value=reserva.get_status_display())
                ws.cell(row=row, column=8, value=reserva.data_reserva.strftime('%d/%m/%Y %H:%M'))
                ws.cell(row=row, column=9, value=reserva.data_expiracao.strftime('%d/%m/%Y %H:%M') if reserva.data_expiracao else 'N/A')
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="reservas.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(self.request, 'Biblioteca openpyxl não está instalada para exportação Excel.')
            return redirect('biblioteca:reserva_list')
    
    def export_pdf(self, queryset):
        try:
            from django.template.loader import render_to_string
            from weasyprint import HTML
            from django.http import HttpResponse
            
            html_string = render_to_string('biblioteca/reserva_pdf.html', {
                'reservas': queryset,
                'data_exportacao': timezone.now()
            })
            
            html = HTML(string=html_string)
            pdf = html.write_pdf()
            
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="reservas.pdf"'
            return response
            
        except ImportError:
            messages.error(self.request, 'Biblioteca WeasyPrint não está instalada para exportação PDF.')
            return redirect('biblioteca:reserva_list')